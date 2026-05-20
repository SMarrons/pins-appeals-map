import logging
import json
import os
import io
import datetime
import requests
import openpyxl
from azure.storage.blob import BlobServiceClient, ContentSettings

import azure.functions as func

# ── Config ────────────────────────────────────────────────────────────────────
PINS_XLSX_URL = (
    "https://assets.publishing.service.gov.uk/media/"
    "697ca4a384f2153b1124525e/Casework_Database_Q3.xlsx"
)
# Fallback: scrape the GOV.UK page to find the latest URL dynamically
PINS_PAGE_URL = "https://www.gov.uk/government/publications/planning-inspectorate-appeals-database"

STORAGE_CONN_STR = os.environ["STORAGE_CONNECTION_STRING"]
CONTAINER_NAME   = os.environ.get("STORAGE_CONTAINER", "pins-data")
BLOB_NAME        = "decisions.json"

# ── LPA centroid coordinates (all ~350 English LPAs) ─────────────────────────
LPA_COORDS = {
    "Adur": [50.8376, -0.2839], "Allerdale": [54.6443, -3.3068],
    "Amber Valley": [53.0178, -1.4597], "Arun": [50.8403, -0.5510],
    "Ashfield": [53.1057, -1.2577], "Ashford": [51.1464, 0.8735],
    "Babergh": [52.0213, 0.9373], "Barking and Dagenham": [51.5607, 0.1557],
    "Barnet": [51.6252, -0.1517], "Barnsley": [53.5527, -1.4797],
    "Barrow-in-Furness": [54.1108, -3.2261], "Basildon": [51.5761, 0.4886],
    "Basingstoke and Deane": [51.2667, -1.0876], "Bassetlaw": [53.3237, -0.9500],
    "Bath and North East Somerset": [51.3781, -2.3597], "Bedford": [52.1357, -0.4680],
    "Bexley": [51.4549, 0.1505], "Birmingham": [52.4862, -1.8904],
    "Blaby": [52.5610, -1.1825], "Blackburn with Darwen": [53.7480, -2.4800],
    "Blackpool": [53.8142, -3.0503], "Bolsover": [53.2288, -1.2910],
    "Bolton": [53.5781, -2.4290], "Boston": [52.9761, -0.0202],
    "Bournemouth, Christchurch and Poole": [50.7192, -1.8808],
    "Bracknell Forest": [51.4154, -0.7486], "Bradford": [53.7950, -1.7594],
    "Braintree": [51.8776, 0.5498], "Breckland": [52.5710, 0.7237],
    "Brentwood": [51.6210, 0.3052], "Brighton and Hove": [50.8225, -0.1372],
    "Bristol, City of": [51.4545, -2.5879], "Broadland": [52.6765, 1.2937],
    "Bromley": [51.4063, 0.0143], "Bromsgrove": [52.3369, -2.0568],
    "Broxbourne": [51.7457, -0.0214], "Broxtowe": [52.9435, -1.2279],
    "Buckinghamshire": [51.8097, -0.8146], "Burnley": [53.7891, -2.2478],
    "Bury": [53.5933, -2.2966], "Calderdale": [53.7249, -1.8658],
    "Cambridge": [52.2053, 0.1218], "Camden": [51.5390, -0.1426],
    "Cannock Chase": [52.6897, -2.0282], "Canterbury": [51.2802, 1.0789],
    "Carlisle": [54.8951, -2.9382], "Castle Point": [51.5559, 0.6974],
    "Central Bedfordshire": [52.0024, -0.4662], "Chelmsford": [51.7361, 0.4798],
    "Cheltenham": [51.8994, -2.0783], "Cherwell": [52.0271, -1.2815],
    "Cheshire East": [53.1615, -2.2185], "Cheshire West and Chester": [53.1912, -2.7330],
    "Chichester": [50.8365, -0.7792], "Charnwood": [52.7270, -1.1866],
    "Chorley": [53.6518, -2.6326], "City of London": [51.5155, -0.0922],
    "Colchester": [51.8959, 0.8919], "Copeland": [54.5296, -3.5026],
    "Cornwall": [50.2660, -5.0527], "Cotswold": [51.7393, -1.8418],
    "Coventry": [52.4068, -1.5197], "Craven": [54.0658, -2.0532],
    "Crawley": [51.1092, -0.1872], "Croydon": [51.3714, -0.0977],
    "Cumberland": [54.6393, -3.0010], "Dacorum": [51.7601, -0.4680],
    "Darlington": [54.5236, -1.5538], "Dartford": [51.4429, 0.2183],
    "Derby": [52.9225, -1.4746], "Derbyshire Dales": [53.0823, -1.6985],
    "Doncaster": [53.5228, -1.1286], "Dover": [51.1279, 1.3134],
    "Dudley": [52.5128, -2.0861], "Durham, County": [54.7753, -1.5849],
    "Ealing": [51.5130, -0.3089], "East Cambridgeshire": [52.3510, 0.2523],
    "East Devon": [50.7400, -3.3067], "East Hampshire": [51.0573, -1.0000],
    "East Hertfordshire": [51.8751, 0.0563], "East Lindsey": [53.1903, 0.1528],
    "East Riding of Yorkshire": [53.8419, -0.4342], "East Staffordshire": [52.8033, -1.6888],
    "East Suffolk": [52.1552, 1.4037], "Eastbourne": [50.7688, 0.2799],
    "Eastleigh": [50.9697, -1.3521], "Eden": [54.6590, -2.7574],
    "Elmbridge": [51.3540, -0.3754], "Enfield": [51.6538, -0.0799],
    "Epping Forest": [51.7026, 0.1102], "Epsom and Ewell": [51.3347, -0.2692],
    "Erewash": [52.8897, -1.2894], "Exeter": [50.7184, -3.5339],
    "Fareham": [50.8524, -1.1784], "Fenland": [52.5340, 0.1009],
    "Folkestone and Hythe": [51.0817, 1.1618], "Forest of Dean": [51.7926, -2.5594],
    "Fylde": [53.7912, -2.8905], "Gateshead": [54.9630, -1.6033],
    "Gedling": [52.9930, -1.0786], "Gloucester": [51.8642, -2.2381],
    "Gosport": [50.7953, -1.1180], "Gravesham": [51.4419, 0.3689],
    "Great Yarmouth": [52.6078, 1.7298], "Greenwich": [51.4934, 0.0098],
    "Guildford": [51.2365, -0.5703], "Hackney": [51.5450, -0.0553],
    "Halton": [53.3687, -2.7265], "Hambleton": [54.2936, -1.4096],
    "Hammersmith and Fulham": [51.4927, -0.2239], "Harborough": [52.4815, -0.9226],
    "Haringey": [51.5906, -0.1114], "Harlow": [51.7624, 0.0946],
    "Harrogate": [54.0033, -1.5412], "Harrow": [51.5836, -0.3464],
    "Hart": [51.2753, -0.9484], "Hartlepool": [54.6910, -1.2126],
    "Hastings": [50.8543, 0.5730], "Havant": [50.8565, -1.0119],
    "Havering": [51.5779, 0.2120], "Herefordshire": [52.0655, -2.7152],
    "Hertsmere": [51.6967, -0.2700], "High Peak": [53.3695, -1.8828],
    "Hillingdon": [51.5441, -0.4760], "Hinckley and Bosworth": [52.5406, -1.3766],
    "Horsham": [51.0624, -0.3225], "Hounslow": [51.4746, -0.3680],
    "Huntingdonshire": [52.3300, -0.1772], "Hyndburn": [53.7580, -2.3888],
    "Ipswich": [52.0567, 1.1482], "Isle of Wight": [50.6938, -1.3047],
    "Isles of Scilly": [49.9255, -6.3161], "Islington": [51.5362, -0.1033],
    "Kensington and Chelsea": [51.4990, -0.1919], "Kettering": [52.3978, -0.7297],
    "King's Lynn and West Norfolk": [52.7520, 0.4000],
    "Kingston upon Hull, City of": [53.7457, -0.3367],
    "Kingston upon Thames": [51.4085, -0.3064], "Kirklees": [53.5935, -1.8011],
    "Knowsley": [53.4547, -2.8523], "Lambeth": [51.4571, -0.1231],
    "Lancaster": [54.0466, -2.8007], "Leeds": [53.8008, -1.5491],
    "Leicester": [52.6369, -1.1398], "Lewes": [50.8745, 0.0115],
    "Lewisham": [51.4415, -0.0117], "Lichfield": [52.6821, -1.8279],
    "Lincoln": [53.2307, -0.5406], "Liverpool": [53.4084, -2.9916],
    "Luton": [51.8787, -0.4200], "Maidstone": [51.2722, 0.5217],
    "Maldon": [51.7321, 0.6753], "Malvern Hills": [52.1191, -2.3150],
    "Manchester": [53.4808, -2.2426], "Mansfield": [53.1430, -1.1945],
    "Medway": [51.3900, 0.5403], "Melton": [52.7631, -0.8860],
    "Mendip": [51.2303, -2.5978], "Merton": [51.4014, -0.1965],
    "Mid Devon": [50.8715, -3.6956], "Mid Suffolk": [52.1893, 1.0102],
    "Mid Sussex": [51.0032, -0.1614], "Middlesbrough": [54.5741, -1.2350],
    "Milton Keynes": [52.0406, -0.7594], "Mole Valley": [51.2293, -0.3277],
    "New Forest": [50.9051, -1.6033], "Newark and Sherwood": [53.0759, -0.8082],
    "Newcastle upon Tyne": [54.9783, -1.6178],
    "Newcastle-under-Lyme": [53.0116, -2.2274],
    "Newham": [51.5255, 0.0352], "North Devon": [51.0508, -3.9940],
    "North East Derbyshire": [53.2885, -1.4470],
    "North East Lincolnshire": [53.5657, -0.0798],
    "North Hertfordshire": [51.9332, -0.2497],
    "North Kesteven": [53.0050, -0.4860], "North Lincolnshire": [53.6039, -0.5530],
    "North Norfolk": [52.9340, 1.0869], "North Somerset": [51.3879, -2.7755],
    "North Tyneside": [55.0182, -1.4860], "North Warwickshire": [52.5594, -1.5846],
    "North West Leicestershire": [52.7340, -1.3632],
    "North Yorkshire": [54.1610, -1.5264], "Northampton": [52.2405, -0.9027],
    "Northumberland": [55.2082, -1.9855], "Norwich": [52.6309, 1.2974],
    "Nottingham": [52.9548, -1.1581], "Nuneaton and Bedworth": [52.5238, -1.4661],
    "Oadby and Wigston": [52.5883, -1.0872], "Oldham": [53.5444, -2.1169],
    "Oxford": [51.7520, -1.2577], "Pendle": [53.8540, -2.2103],
    "Peterborough": [52.5696, -0.2405], "Plymouth": [50.3755, -4.1427],
    "Portsmouth": [50.8198, -1.0880], "Preston": [53.7632, -2.7031],
    "Reading": [51.4543, -0.9781], "Redbridge": [51.5590, 0.0741],
    "Redcar and Cleveland": [54.6019, -1.0819], "Redditch": [52.3073, -1.9462],
    "Reigate and Banstead": [51.2369, -0.1623], "Ribble Valley": [53.8740, -2.4340],
    "Richmond upon Thames": [51.4479, -0.3260],
    "Richmondshire": [54.3962, -1.7435], "Rochdale": [53.6097, -2.1561],
    "Rochford": [51.5816, 0.7072], "Rossendale": [53.7042, -2.2784],
    "Rother": [50.9169, 0.5781], "Rotherham": [53.4302, -1.3568],
    "Rugby": [52.3701, -1.2643], "Runnymede": [51.3967, -0.5316],
    "Rushcliffe": [52.8735, -1.0680], "Rushmoor": [51.2942, -0.7562],
    "Rutland": [52.6580, -0.6394], "Ryedale": [54.1383, -0.9060],
    "Salford": [53.4875, -2.2901], "Sandwell": [52.5062, -2.0126],
    "Scarborough": [54.2766, -0.4066], "Sedgemoor": [51.1278, -2.9631],
    "Sefton": [53.5436, -2.9767], "Selby": [53.7827, -1.0692],
    "Sevenoaks": [51.2858, 0.1871], "Sheffield": [53.3811, -1.4701],
    "Shepway": [51.0817, 1.1618], "Shropshire": [52.6781, -2.7573],
    "Slough": [51.5105, -0.5950], "Solihull": [52.4130, -1.7774],
    "Somerset": [51.1050, -2.9861], "Somerset West and Taunton": [51.0193, -3.1006],
    "South Cambridgeshire": [52.1790, 0.0733],
    "South Derbyshire": [52.8038, -1.5590],
    "South Gloucestershire": [51.5274, -2.4778],
    "South Hams": [50.3600, -3.8106], "South Holland": [52.7798, -0.1122],
    "South Kesteven": [52.8294, -0.6357], "South Lakeland": [54.3269, -2.7340],
    "South Norfolk": [52.4510, 1.2281],
    "South Northamptonshire": [52.2098, -1.1153],
    "South Oxfordshire": [51.5975, -1.1335],
    "South Ribble": [53.7429, -2.6762], "South Somerset": [51.0131, -2.7086],
    "South Staffordshire": [52.5848, -2.1948],
    "South Tyneside": [54.9632, -1.4420], "Southampton": [50.9097, -1.4044],
    "Southend-on-Sea": [51.5384, 0.7148], "Southwark": [51.5035, -0.0804],
    "Spelthorne": [51.4161, -0.5011], "St Albans": [51.7454, -0.3365],
    "St. Helens": [53.4550, -2.7367], "Stafford": [52.8064, -2.1169],
    "Staffordshire Moorlands": [53.0622, -1.9812],
    "Stevenage": [51.9024, -0.2019], "Stockport": [53.4106, -2.1575],
    "Stockton-on-Tees": [54.5700, -1.3173], "Stoke-on-Trent": [53.0027, -2.1794],
    "Stratford-on-Avon": [52.1916, -1.7083], "Stroud": [51.7448, -2.2169],
    "Suffolk Coastal": [52.0716, 1.4900], "Sunderland": [54.9058, -1.3817],
    "Surrey Heath": [51.3217, -0.7369], "Sutton": [51.3618, -0.1945],
    "Swale": [51.3368, 0.7394], "Tameside": [53.4806, -2.0820],
    "Tamworth": [52.6390, -1.6957], "Tandridge": [51.2371, -0.0680],
    "Teignbridge": [50.5200, -3.6163], "Telford and Wrekin": [52.6955, -2.4515],
    "Tendring": [51.8483, 1.1573], "Test Valley": [51.0958, -1.5085],
    "Tewkesbury": [51.9919, -2.1604], "Thanet": [51.3597, 1.3925],
    "Three Rivers": [51.6663, -0.4261], "Thurrock": [51.4929, 0.3530],
    "Tonbridge and Malling": [51.1966, 0.2900], "Torbay": [50.4619, -3.5253],
    "Torridge": [50.9870, -4.1566], "Tower Hamlets": [51.5099, -0.0059],
    "Trafford": [53.4267, -2.3296], "Tunbridge Wells": [51.1323, 0.2630],
    "Uttlesford": [51.9679, 0.3443], "Vale of White Horse": [51.6659, -1.3773],
    "Wakefield": [53.6833, -1.4977], "Walsall": [52.5860, -1.9826],
    "Waltham Forest": [51.5908, -0.0134], "Wandsworth": [51.4571, -0.1923],
    "Warrington": [53.3900, -2.5970], "Warwick": [52.2819, -1.5860],
    "Watford": [51.6565, -0.3964], "Waverley": [51.1842, -0.5912],
    "Wealden": [50.9300, 0.2300], "Wellingborough": [52.3018, -0.6888],
    "Welwyn Hatfield": [51.7623, -0.2200], "West Berkshire": [51.4403, -1.3028],
    "West Devon": [50.6094, -4.1069], "West Lancashire": [53.6073, -2.7764],
    "West Lindsey": [53.3702, -0.5343],
    "West Northamptonshire": [52.2396, -0.8960],
    "West Oxfordshire": [51.7861, -1.5607], "West Suffolk": [52.2463, 0.6878],
    "Westminster": [51.4975, -0.1357],
    "Westmorland and Furness": [54.4609, -2.7402],
    "Wigan": [53.5387, -2.6302], "Winchester": [51.0632, -1.3078],
    "Windsor and Maidenhead": [51.5081, -0.7198], "Wirral": [53.3726, -3.0735],
    "Woking": [51.3168, -0.5600], "Wokingham": [51.4112, -0.8347],
    "Wolverhampton": [52.5862, -2.1293], "Worcester": [52.1936, -2.2219],
    "Worthing": [50.8179, -0.3731], "Wychavon": [52.1483, -2.0386],
    "Wyre": [53.8574, -2.9681], "Wyre Forest": [52.3836, -2.2259],
    "York": [53.9590, -1.0815],
}

# ── Column name aliases (PINS column headers vary between releases) ───────────
# Maps our internal key → list of possible header strings to try in order
COLUMN_MAP = {
    "reference":       ["Case Number"],
    "lpa":             ["Local Planning Authority", "LPA Name", "LPA", "LOCAL_PLANNING_AUTHORITY"],
    "outcome":         ["Decision", "Appeal Decision", "Outcome", "DECISION"],
    "appeal_type":     ["Procedure", "Appeal Type", "Type of Appeal", "PROCEDURE"],
    "dev_type":        ["Development Type", "Type of Development", "DEV_TYPE"],
    "decision_date":   ["Decision Date", "Date of Decision", "DECISION_DATE"],
    "inspector":       ["Inspector", "Appointed Inspector", "Inspector Name", "INSPECTOR"],
    "address":         ["Site Address", "Address", "Location", "SITE_ADDRESS"],
    "appeal_start":    ["Start Date", "Appeal Start Date", "START_DATE"],
    "questionnaire":   ["Questionnaire Due", "Questionnaire Date"],
}

# Standardise common outcome variants to our four canonical values
OUTCOME_ALIASES = {
    "allowed":               "Allowed",
    "dismissed":             "Dismissed",
    "split decision":        "Split decision",
    "allowed in part":       "Allowed in part",
    "invalid":               "Invalid",
    "withdrawn":             "Withdrawn",
    "transferred":           "Transferred",
}

def normalise_outcome(raw: str) -> str:
    if not raw:
        return "Unknown"
    return OUTCOME_ALIASES.get(raw.strip().lower(), raw.strip().title())

def find_col(headers: list, key: str) -> int | None:
    """Return the column index for the first matching alias, or None."""
    for alias in COLUMN_MAP.get(key, []):
        try:
            return headers.index(alias)
        except ValueError:
            pass
    # Fallback: case-insensitive partial match
    lower_aliases = [a.lower() for a in COLUMN_MAP.get(key, [])]
    for i, h in enumerate(headers):
        if any(a in h.lower() for a in lower_aliases):
            return i
    return None

def geocode(lpa_name: str) -> tuple[float, float] | None:
    """Return (lat, lng) for an LPA name using fuzzy matching."""
    if not lpa_name:
        return None
    # Exact match
    coords = LPA_COORDS.get(lpa_name.strip())
    if coords:
        return coords
    # Partial match (handles "City of Bristol" vs "Bristol, City of" etc.)
    lpa_lower = lpa_name.lower()
    for k, v in LPA_COORDS.items():
        if k.lower() in lpa_lower or lpa_lower in k.lower():
            return v
    return None

def get_latest_xlsx_url() -> str:
    """Scrape the GOV.UK page to find the current .xlsx download URL."""
    try:
        resp = requests.get(PINS_PAGE_URL, timeout=30)
        resp.raise_for_status()
        # Find first .xlsx link in the page
        import re
        matches = re.findall(r'href="(https://assets\.publishing\.service\.gov\.uk[^"]+\.xlsx)"', resp.text)
        if matches:
            logging.info(f"Found XLSX URL from page: {matches[0]}")
            return matches[0]
    except Exception as e:
        logging.warning(f"Could not scrape page for URL: {e}")
    return PINS_XLSX_URL  # fall back to hardcoded URL

def download_xlsx(url: str) -> bytes:
    logging.info(f"Downloading PINS data from {url}")
    resp = requests.get(url, timeout=120, stream=True)
    resp.raise_for_status()
    data = resp.content
    logging.info(f"Downloaded {len(data):,} bytes")
    return data

def parse_xlsx(data: bytes) -> list[dict]:
    """Parse the PINS Excel file into a list of decision dicts."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    # PINS workbook has multiple sheets; find the Appeals one
    sheet = None
    for name in wb.sheetnames:
        lower = name.lower()

        if "data" in lower:
            sheet = wb[name]
            break

        if "appeal" in lower or "casework" in lower:
            sheet = wb[name]

    if sheet is None:
        sheet = wb.active

    print("Using sheet:", sheet.title)
    print("Workbook sheets:", wb.sheetnames)

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Sheet is empty")

    # Find the header row (first row containing "Reference" or "LPA")
    header_row_idx = 0
    for i, row in enumerate(rows[:10]):
        cells = [str(c).strip() if c else "" for c in row]
        if any("reference" in c.lower() or "lpa" in c.lower() or "decision" in c.lower() for c in cells):
            header_row_idx = i
            break

    headers = [str(c).strip() if c else "" for c in rows[header_row_idx]]
    print("Headers found:", headers)
    logging.info(f"Headers found: {headers}")

    # Resolve column indices
    cols = {key: find_col(headers, key) for key in COLUMN_MAP}
    print("Column mapping:", {k: v for k, v in cols.items() if v is not None})

    records = []
    skipped = 0

    for row in rows[header_row_idx + 1:]:
        if not any(row):
            continue

        def cell(key: str) -> str:
            idx = cols.get(key)
            if idx is None or idx >= len(row):
                return ""
            val = row[idx]
            if val is None:
                return ""
            if isinstance(val, datetime.datetime):
                return val.strftime("%Y-%m-%d")
            if isinstance(val, datetime.date):
                return val.isoformat()
            return str(val).strip()

        lpa_name = cell("lpa")
        coords = geocode(lpa_name)

        if not coords:
            skipped += 1
            continue

        outcome_raw = cell("outcome")
        outcome = normalise_outcome(outcome_raw)

        # Skip non-decision rows (blank outcomes, headers repeated mid-sheet)
        if not outcome or outcome in ("", "Decision"):
            skipped += 1
            continue

        reference = cell("reference")
        initials = "".join(w[0] for w in lpa_name.split() if w)[:2].upper()

        records.append({
            "id":              reference,
            "lpa":             lpa_name,
            "lat":             round(coords[0], 5),
            "lng":             round(coords[1], 5),
            "outcome":         outcome,
            "appealType":      cell("appeal_type"),
            "developmentType": cell("dev_type"),
            "decisionDate":    cell("decision_date"),
            "inspector":       cell("inspector"),
            "address":         cell("address"),
            "initials":        initials,
        })

    logging.info(f"Parsed {len(records):,} records ({skipped:,} skipped/ungeocodeable)")
    return records

def upload_json(records: list[dict]) -> str:
    """Write decisions.json to Azure Blob Storage and return the public URL."""
    payload = {
        "generatedAt": datetime.datetime.utcnow().isoformat() + "Z",
        "count":       len(records),
        "records":     records,
    }
    json_bytes = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")

    client = BlobServiceClient.from_connection_string(STORAGE_CONN_STR)
    container = client.get_container_client(CONTAINER_NAME)

    # Ensure container exists with public blob access
    try:
        container.create_container(public_access="blob")
    except Exception:
        pass  # already exists

    blob = container.get_blob_client(BLOB_NAME)
    blob.upload_blob(
        json_bytes,
        overwrite=True,
        content_settings=ContentSettings(
            content_type="application/json",
            cache_control="public, max-age=3600",
        ),
    )

    url = blob.url
    logging.info(f"Uploaded {len(json_bytes):,} bytes → {url}")
    return url

# ── Azure Function entry point ────────────────────────────────────────────────
def main(mytimer: func.TimerRequest) -> None:
    utc_now = datetime.datetime.utcnow()
    logging.info(f"PINS downloader triggered at {utc_now.isoformat()}")

    if mytimer.past_due:
        logging.warning("Timer is past due — running anyway")

    try:
        url   = get_latest_xlsx_url()
        data  = download_xlsx(url)
        recs  = parse_xlsx(data)
        blob_url = upload_json(recs)
        logging.info(f"SUCCESS — {len(recs):,} decisions written to {blob_url}")
    except Exception as exc:
        logging.exception(f"PINS downloader failed: {exc}")
        raise
